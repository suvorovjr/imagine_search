import openpyxl
from pathlib import Path

ABSOLUTH_PATH = Path(__file__).parent
IMAGES = ABSOLUTH_PATH.joinpath("imagine") # укажи свой путь к файлам
images_names = list(IMAGES.iterdir())
workbook = openpyxl.load_workbook("articles.xlsx")
sheet = workbook.active
rows = sheet.max_row
articles = []
article_image_mapping = {}
for i in range(rows):
    articl = sheet.cell(row=i + 10, column=4).value
    articles.append(articl)

article_image_mapping = {}

article_row_mapping = {}
for row_num, row in enumerate(sheet.iter_rows(min_row=10, max_row=rows, min_col=4, max_col=4, values_only=True), start=10):
    article = str(row[0])
    article_row_mapping[article] = row_num

for name_image in images_names:
    image_path = Path(name_image)
    image_name = image_path.stem.lower()
    image_name = "".join(char for char in image_name if char.isalpha() or char.isdigit())

    for excel_article in articles:
        if not isinstance(excel_article, int):

            article_to_compare = str(excel_article)
            if article_to_compare.endswith("23л") or article_to_compare.endswith("23з"):
                article_to_compare = article_to_compare[:-3]

            article_to_compare = "".join(char for char in article_to_compare if char.isalpha() or char.isdigit()).lower()

            if article_to_compare in image_name:
                if excel_article not in article_image_mapping:
                    article_image_mapping[excel_article] = []
                article_image_mapping[excel_article].append(image_path)
for article, image_paths in article_image_mapping.items():
    for i, image_path in enumerate(image_paths):
        new_name = f"{article}-{i + 1}.jpg"
        target_path = IMAGES.joinpath(new_name)
        try:
            image_path.rename(target_path)
        except FileNotFoundError:
            pass

for article, row_num in article_row_mapping.items():
    cell = sheet.cell(row=row_num, column=7)
    image_names = []
    for i, image_path in enumerate(article_image_mapping.get(article, [])):
        new_name = f"{article}-{i + 1}.jpg"
        image_names.append(f"https://imperiya-detstva.ru/upload/SW23/{new_name}/")
    cell.value = "\n".join(image_names)

workbook.save("articles.xlsx")